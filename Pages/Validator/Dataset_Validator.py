import streamlit as st
import os
from io import BytesIO
import pandas as pd
from pandas import read_csv, ExcelWriter, DataFrame, read_table
from os import chdir, remove
from os.path import dirname, basename
from zipfile import ZipFile, ZIP_DEFLATED
from datetime import time
from datetime import datetime
import re
from dateutil.parser import parse
from openpyxl import load_workbook


class Instructions:
    instructions = 'Upload the Dataset xlsx and run the script to download a validator excel report file'
    link = 'https://optibus.atlassian.net/wiki/spaces/OP/pages/3117940752/JDF+Converting+Scripts#JDF-to-Dataset'

class Description:
    title = "Dataset Validator"
    description = "This is a script that validating your Dataset file, the output is an Excel file the lists all the error/warnings"
    icon="https://pnx-assets-prod.s3.amazonaws.com/2023-05/netiq_validator.png"
    author = 'Lior Zacks'

def run():
    
    logger = st.expander('logging outputs for debugging')

    def main():
        uploaded_file = st.file_uploader("Select Dataset Excel file", type=['xlsx'])
        if uploaded_file is not None:
            time_0 = time()
            output_name ='Dataset_validator_Output'
            dict={}
            VehicleTypes = validate_VehicleTypes(uploaded_file)
            dict['VehicleTypes']=VehicleTypes
            Places = validate_Places(uploaded_file)
            dict['Places']=Places
            StopTimes = validate_StopTimes(uploaded_file)
            dict['StopTimes']=StopTimes
            Trips = validate_trips(uploaded_file)
            dict['Trips']=Trips
            Trips_with_StopTimes = validate_trips_with_stoptimes(uploaded_file)
            dict['Trips_with_StopTimes']=Trips_with_StopTimes
            trips_with_veihcles = validate_trips_with_veihcles(uploaded_file)
            dict['trips_with_veihcles']=trips_with_veihcles
            trips_with_places = validate_trips_with_places(uploaded_file)
            dict['trips_with_places'] = trips_with_places
            general_notes=general()
            dict['general_notes']=general_notes
            output=write_excel(dict)
            st.download_button("Download Excel File", output, output_name + '.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def validate_VehicleTypes(excel_sheet):
        wb = load_workbook(excel_sheet, read_only=True)  # open an Excel file and return a workbook
        if 'VehicleTypes' not in wb.sheetnames:
            data = {'VehicleTypes': [
                'VehicleTypes sheet not exists']}
            df = pd.DataFrame(data)
            return df
        df = pd.read_excel(excel_sheet, sheet_name='VehicleTypes')
        errors = []
        # Check column names
        id_col = df.columns[0]
        short_name_col = df.columns[1]
        if id_col not in ["Vehicle Type Id", "Id", "קוד"]:
            errors.append((id_col, f'Unexpected column name: {id_col}'))
        if short_name_col not in ["Vehicle Type Name", "name", "Description", "שם", "סוח", "short_name"]:
            errors.append((short_name_col, f'Unexpected column name: {short_name_col}'))
    
        # Check at least one non-empty row in column 0
        if df[id_col].isnull().all():
            errors.append((id_col, 'All values are missing in column 0'))
    
        # Check for special characters in column 0
        special_chars = df[id_col].apply(lambda x: bool(re.search(r'[^a-zA-Z0-9_ ]', str(x))))
        if special_chars.any():
            errors.append((id_col, "Warning: Special character found, it may cause errors"))
    
        # Check for duplicates
        if df[id_col].duplicated().any():
            errors.append((id_col, 'Duplicates found in column 0'))
        if df[short_name_col].duplicated().any():
            errors.append((short_name_col, 'Duplicates found in column 1'))
    
        # Check for leading/trailing spaces
        df[id_col] = df[id_col].astype(str)
        print(df[id_col].dtype)
        st.write(df[id_col].dtype)
        if df[id_col].str.contains(r'^\s|\s$', na=False).any():
            errors.append((id_col, 'Found leading/trailing spaces in column 0'))
        if df[short_name_col].str.contains(r'^\s|\s$', na=False).any():
            errors.append((short_name_col, 'Found leading/trailing spaces in column 1'))
    
        return pd.DataFrame(errors, columns=['Column', 'Error'])

    def validate_Places(excel_sheett):
        wb = load_workbook(excel_sheett, read_only=True)  # open an Excel file and return a workbook
        if 'Places' not in wb.sheetnames:
            data = {'Places': [
                'Places sheet not exists']}
            df = pd.DataFrame(data)
            return df
        excel_sheet = pd.read_excel(excel_sheett, sheet_name='Places')
        errors = []
    
        # Define column names and column mappings
        column_mappings = {
            'id': ["Stop Id", "Id", "קוד"],
            'description': ["description","Stop Name", "name", "Description","שם"],
            'latitude': ["Latitude", "stop_lat", "קו רוחב", "רוחב"],
            'longitude': ["Longitude", "stop_lon", "קו אורך", "אורך"],
            'type': ["depot", "Depot", "Type", "חניון", "סוג"],
            'place': ["Place", "Places", "מקום"]
        }
    
        # Expected positions for mandatory columns
        mandatory_positions = {
            'id': [0],
            'description': [1],
            'latitude': [2, 3, 4],
            'longitude': [3, 4, 5],
            'type': [4, 5, 6]
        }
    
        # Optional positions
        optional_positions = {
            'place': [2]
        }
    
        # Assign columns based on column_mappings
        df_columns = {}
        for key, value in column_mappings.items():
            for i, column in enumerate(excel_sheet.columns):
                if column in value:
                    # Check if the column is in its expected position
                    if key in mandatory_positions and i not in mandatory_positions[key]:
                        errors.append((column, 'Column in unexpected position', 'N/A'))
                    if key in optional_positions and i not in optional_positions[key]:
                        errors.append((column, 'Column in unexpected position', 'N/A'))
                    df_columns[key] = column
    
        # Check for existence of mandatory columns
        for key in mandatory_positions.keys():
            if key not in df_columns:
                errors.append((key, 'Mandatory column is missing', 'N/A'))
    
        # If there are missing mandatory columns, return the error DataFrame
        if len(errors) > 0:
            return pd.DataFrame(errors, columns=['Column', 'Error', 'Row'])
    
        # Perform validations on each mandatory column except 'latitude', 'longitude', 'type'
        for key in ['id']:
            # Check for duplicates
            if excel_sheet[df_columns[key]].duplicated().any():
                duplicates = excel_sheet[excel_sheet[df_columns[key]].duplicated(keep=False)]
                for index, row in duplicates.iterrows():
                    errors.append((df_columns[key], 'Duplicate value found', index+2))
    
            # Check for empty values
            if excel_sheet[df_columns[key]].isnull().any():
                empties = excel_sheet[excel_sheet[df_columns[key]].isnull()]
                for index, row in empties.iterrows():
                    errors.append((df_columns[key], 'Empty value found', index+2))
    
            # Check for special characters in 'id'
            if key == 'id':
                special_chars = excel_sheet[df_columns['id']].apply(lambda x: bool(re.search(r'[^a-zA-Z0-9_ ]', str(x))))
                if special_chars.any():
                    special_chars = excel_sheet[special_chars]
                    for index, row in special_chars.iterrows():
                        errors.append((df_columns['id'], 'Warning:Special character found, it may cause errors', index+2))
    
        # Check for empty values
        for key in ['latitude', 'longitude']:
            if excel_sheet[df_columns[key]].isnull().any():
                empties = excel_sheet[excel_sheet[df_columns[key]].isnull()]
                for index, row in empties.iterrows():
                    errors.append((df_columns[key], 'Empty value found', index+2))
        lat = pd.to_numeric(excel_sheet[df_columns['latitude']], errors='coerce')
        lon = pd.to_numeric(excel_sheet[df_columns['longitude']], errors='coerce')
    
    
        # If 'place' exists, check its values
        if 'place' in df_columns:
            # Check that values in 'place' column are not equal to 'id' or 'description' values
            if (excel_sheet[df_columns['place']] == excel_sheet[df_columns['id']]).any() or (excel_sheet[df_columns['place']] == excel_sheet[df_columns['description']]).any():
                wrong_values = excel_sheet[(excel_sheet[df_columns['place']] == excel_sheet[df_columns['id']]) | (excel_sheet[df_columns['place']] == excel_sheet[df_columns['description']])]
                for index, row in wrong_values.iterrows():
                    errors.append((df_columns['place'], '"Place" values are equal to "Id" or "Description" values', index+2))
    
            # Check that 'place' values don't appear in 'id' or 'description' columns
            place_values = excel_sheet[df_columns['place']].unique()
            for value in place_values:
                if (excel_sheet[df_columns['id']] == value).any() or (excel_sheet[df_columns['description']] == value).any():
                    wrong_values = excel_sheet[(excel_sheet[df_columns['id']] == value) | (excel_sheet[df_columns['description']] == value)]
                    for index, row in wrong_values.iterrows():
                        errors.append((df_columns['place'], f'"Place" value "{value}" found in "Id" or "Description"', index+2))
    
        return pd.DataFrame(errors, columns=['Column', 'Error', 'Row'])
    
    def parse_time(x):
        if pd.isnull(x):
            return None
        for fmt in ('%H:%M', '%H:%M:%S', '%H:%M:%S.%f'):
            try:
                return datetime.strptime(x, fmt).time()
            except ValueError:
                pass
        raise ValueError('No valid time format found')
    
    def validate_time_format(time_str):
        try:
            datetime.strptime(time_str, "%H:%M")
            return True
        except ValueError:
            try:
                datetime.strptime(time_str, "%H:%M:%S")
                return True
            except ValueError:
                try:
                    time_obj = datetime.strptime(time_str, "%H:%M:%S.%f")
                    if time_obj.hour <= 36 or (time_obj.hour == 37 and time_obj.minute == 0 and time_obj.second == 0):
                        return True
                    else:
                        return False
                except ValueError:
                    return False
    
    def validate_StopTimes(dff):
        wb = load_workbook(dff, read_only=True)  # open an Excel file and return a workbook
        if 'StopTimes' not in wb.sheetnames:
            data = {'StopTimes': [
                'StopTimes sheet not exists']}
            df = pd.DataFrame(data)
            return df
        df = pd.read_excel(dff, sheet_name='StopTimes')
        errors = []  # to store the errors
    
        # Test 1: Check column index 0
        column_name_0 = df.columns[0]
        allowed_names_0 = ["Trip Id", "Id", "מספר סידורי", "Trip ID"]
        if column_name_0 not in allowed_names_0:
            errors.append([column_name_0, f"Name must be one of: 'Trip Id', 'Id', 'מספר סידורי', 'Trip ID'", None])
    
        empty_values_0 = df[column_name_0].isnull() | df[column_name_0].astype(str).str.strip().eq('')
        if empty_values_0.any():
            empty_rows_0 = df.index[empty_values_0].tolist()
            for row in empty_rows_0:
                errors.append([column_name_0, "Empty value found", row + 2])
    
        # Test 2: Check column index 1
        column_name_1 = df.columns[1]
        allowed_names_1 = ["Point Time", "Time"]
        if column_name_1 not in allowed_names_1:
            errors.append([column_name_1, f"Name must be one of: 'Point Time', 'Time'", None])
    
        empty_values_1 = df[column_name_1].isnull() | df[column_name_1].astype(str).str.strip().eq('')
        if empty_values_1.any():
            empty_rows_1 = df.index[empty_values_1].tolist()
            for row in empty_rows_1:
                errors.append([column_name_1, "Empty value found", row + 2])
    
        # Test 3: Check column index 2
        column_name_2 = df.columns[2]
        allowed_names_2 = ["Point ID", "Point Id", "Stop ID", "Stop Id"]
        if column_name_2 not in allowed_names_2:
            errors.append([column_name_2, f"Name must be one of: 'Point ID', 'Point Id', 'Stop ID', 'Stop Id'", None])
    
        empty_values_2 = df[column_name_2].isnull() | df[column_name_2].astype(str).str.strip().eq('')
        if empty_values_2.any():
            empty_rows_2 = df.index[empty_values_2].tolist()
            for row in empty_rows_2:
                errors.append([column_name_2, "Empty value found", row + 2])
    
        # Test 4: Check "Time Point" column
        column_name_3 = "Time Point"
        allowed_names_3 = ["Time Point", "Timepoint", "Trace Point"]
        if column_name_3 not in df.columns:
            errors.append([column_name_3, f"'{column_name_3}' column is missing", None])
        else:
            empty_values_3 = df[column_name_3].isnull() | df[column_name_3].astype(str).str.strip().eq('')
            if empty_values_3.any():
                empty_rows_3 = df.index[empty_values_3].tolist()
                for row in empty_rows_3:
                    errors.append([column_name_3, "Empty value found", row + 2])
    
        # Test 5: Check "Sequence" column
        column_name_4 = None
        allowed_names_4 = ["Sequence", "sequence"]
        for name in allowed_names_4:
            if name in df.columns:
                column_name_4 = name
                break
    
        if column_name_4 is None:
            errors.append(["Sequence", "'Sequence' column is missing", None])
        else:
            sequence_column = df[column_name_4]
            empty_values_4 = sequence_column.isnull() | sequence_column.astype(str).str.strip().eq('')
            if empty_values_4.any():
                empty_rows_4 = df.index[empty_values_4].tolist()
                for row in empty_rows_4:
                    errors.append([column_name_4, "Empty value found", row + 2])
    
            sequence_column = pd.to_numeric(sequence_column, errors='coerce')
            df[column_name_4] = sequence_column  # Update the sequence column in the dataframe
    
            # Check if the two rows with the lowest sequence column values within each trip have different values in column index 2
            trips = df[column_name_0].unique()
            for trip_id in trips:
                trip_df = df[df[column_name_0] == trip_id]
                min_sequence_values = trip_df[column_name_4].nsmallest(2)
                if len(min_sequence_values) > 1 and min_sequence_values.iloc[0] == min_sequence_values.iloc[1]:
                    trip_df_min_sequence_rows = trip_df[trip_df[column_name_4].isin(min_sequence_values)]
                    unique_values_in_column_2 = trip_df_min_sequence_rows[column_name_2].nunique()
                    if unique_values_in_column_2 == 1:
                        errors.append([column_name_2,
                                       "Same value in column index 2 for rows with the two lowest sequence values in the same Trip Id",
                                       None])
    
        # Test 6: Check "Distance" column
        column_name_5 = None
        allowed_names_5 = ["Distance", "distance"]
        for name in allowed_names_5:
            if name in df.columns:
                column_name_5 = name
                break
    
        if column_name_5 is not None:
            distance_column = df[column_name_5]
            empty_values_5 = distance_column.isnull() | distance_column.astype(str).str.strip().eq('')
            if empty_values_5.any():
                empty_rows_5 = df.index[empty_values_5].tolist()
                for row in empty_rows_5:
                    errors.append([column_name_5, "Empty value found", row + 2])
    
            # Check if the distance column value is equal to 0 for the row with the lowest sequence in each trip
            for trip_id in trips:
                trip_df = df[df[column_name_0] == trip_id]
                min_sequence_value = trip_df[column_name_4].min()
                distance_value = trip_df.loc[trip_df[column_name_4] == min_sequence_value, column_name_5].values
                if ((len(distance_value) > 0) and (pd.notnull(distance_value[0])) and (distance_value[0] != 0 )):
                    errors.append([column_name_5,
                                   f"Distance value is not equal to 0 for the row with the lowest sequence in Trip Id {trip_id}",
                                   None])
    
        # Test 7: Check for leading empty spaces in all columns
        for column in df.columns:
            leading_spaces = df[column].astype(str).str.startswith(' ')
            if leading_spaces.any():
                leading_spaces_rows = df.index[leading_spaces].tolist()
                for row in leading_spaces_rows:
                    errors.append([column, "Leading spaces found", row + 2])
    
        # Test 8: Check for duplicate rows
        duplicate_rows = df.duplicated()
        if duplicate_rows.any():
            duplicate_indices = df.index[duplicate_rows].tolist()
            for index in duplicate_indices:
                errors.append(["Duplicate Rows", "Duplicate row found", index + 2])
        return pd.DataFrame(errors, columns=["Column Name", "Error", "Row Number"])
    
    def time_to_minutes(t):
        return t.hour * 60 + t.minute
    
    def validate_trips(dff):
        wb = load_workbook(dff, read_only=True)  # open an Excel file and return a workbook
        if 'Trips' not in wb.sheetnames:
            data = {'Trips': [
                'Trips sheet not exists']}
            df = pd.DataFrame(data)
            return df
        df = pd.read_excel(dff, sheet_name='Trips')
        # Create a new dataframe to store the error messages
        error_df = pd.DataFrame(columns=["Column", "Row", "Error_Message"])
        next_index = 0  # index for the next error to be inserted into error_df
    
        # Define a dictionary with column indices and their expected column names
        expected_column_names = {
            0: ["Trip Id", "Id", "מספר סידורי"],
            1: ["Region", "אשכול", "Area"],
            2: ["Catalog Number", "מק\"ט"],
            3: ["Sign", "Route", "קו", "Route Number"],
            4: ["Direction", "כיוון"],
            5: ["Alternative", "חלופה"],
            6: ["Origin Stop Id", "קוד מוצא", "Origin Station ID", "Origin Stop","Origin Stop id"],
            7: ["Destination Stop Id", "קוד יעד", "Destination Station ID", "Destination Stop","Destination Stop id"],
            8: ["Next Day", "Depart next day", "day offset", "Day Offset"],
            9: ["Departure", "שעת יציאה", "departure time"],
            10: ["Arrival", "שעת הגעה", "arrival time"],
            11: ["Vehicle Type Ids", "Vehicle Type Id", "קוד סוג רכב"],
            12: ["Distance", "קילומטרז"],
            13: ["Existing", "סידור", "Schedule"],
            14: ["Custom", "מותאם אישית"],
            15: ["Days", "יום", "יומים", "Day"],
        }
    
        # Define a dictionary with column indices and their expected values
        expected_column_values = {
            8: ["Yes", "No","1", "0", 1, 0]
        }
        # Define a list with column names to be checked for deletion
        columns_to_delete = ["Service Groups", "service groups", "Service IDs", "service ids", "Service Name",
                             "Service name", "service Name"]
    
        # Check if any of the columns to be deleted exist in the DataFrame
        for column_name in columns_to_delete:
            if column_name in df.columns:
                error_df.loc[next_index] = [column_name, None, "Suggestion: delete this column."]
                next_index += 1
    
        # Loop through each column and its expected names, and perform checks
        for column_index, expected_names in expected_column_names.items():
            column_name = df.columns[column_index]
    
            # Check if column name is one of the expected names
            if column_name not in expected_names:
                error_df.loc[next_index] = [column_index, None, f"Column name should be one of {expected_names}."]
                next_index += 1
    
            # Check for empty values in columns 0, 4, 6, 7, 9, 10, 11, 12, and 15
            if column_index in {0, 4, 6, 7, 9, 10, 11, 12, 15} and df[column_name].isnull().any():
                error_rows = df[df[column_name].isnull()].index.tolist()
                for row in error_rows:
                    error_df.loc[next_index] = [column_index, row+2, "Value can't be empty."]
                    next_index += 1
    
            # Check for specific column value constraints
            if column_index in expected_column_values:
                expected_values = expected_column_values[column_index]
                if not df[column_name].isin(expected_values).all():
                    error_rows = df[~df[column_name].isin(expected_values)].index.tolist()
                    for row in error_rows:
                        error_df.loc[next_index] = [column_index, row+2, f"Invalid value. Value should be one of {expected_values}."]
                        next_index += 1
    
            if column_index == 0:  # Check for duplicate values
                if df[column_name].duplicated().any():
                    error_rows = df[df[column_name].duplicated()].index.tolist()
                    for row in error_rows:
                        error_df.loc[next_index] = [column_index, row+2, "Duplicate value detected."]
                        next_index += 1
    
            if column_index == 0:  # Check for special characters
                df[column_name] = df[column_name].astype(str)
                if df[column_name].str.contains('[^\w\s]').any():
                    error_rows = df[df[column_name].str.contains('[^\w\s]')].index.tolist()
                    for row in error_rows:
                        error_df.loc[next_index] = [column_index, row+2, "'Warning:Special character found, it may cause errors'"]
                        next_index += 1
    
            # Check that the hour value in column 9 is lower than the hour value in column 10
            if column_index == 9:
                departure_column = df[column_name].apply(
                    lambda x: time_to_minutes(x) if pd.notnull(x) and isinstance(x, time) else None)
                arrival_column = df[df.columns[10]].apply(
                    lambda x: time_to_minutes(x) if pd.notnull(x) and isinstance(x, time) else None)
    
    
                # Adjusting minutes if they represent 0 hour (midnight)
                departure_column = departure_column.apply(lambda x: x if x != 0 else 24 * 60)
                arrival_column = arrival_column.apply(lambda x: x if x != 0 else 24 * 60)
    
                # Add a day offset to arrival times that are less than departure times
                for i in range(len(arrival_column)):
                    if arrival_column.iloc[i] < departure_column.iloc[i]:
                        arrival_column.iloc[i] += 24 * 60
    
                # Getting rows where departure time is larger than or equal to arrival time
                error_rows = df[
                    (departure_column >= arrival_column) | (departure_column > 36 * 60) | (
                                arrival_column > 36 * 60)].index.tolist()
                for row in error_rows:
                    error_df.loc[next_index] = [column_index, row+2,
                                                "Departure time should be less than arrival time and should be less than or equal to 36 hours."]
                    next_index += 1
            if column_index == 5:
                filtered_df = df[(df[column_name] != "-") & df[column_name].notna()]
    
                # For these rows, check if column index 2 equals column index 3 and Route Id column is empty or doesn't exist
                for index, row in filtered_df.iterrows():
                    if not (pd.isna(row[df.columns[2]]) and pd.isna(row[df.columns[3]])):
                        if row[df.columns[2]] != row[df.columns[3]]:
                            error_df.loc[next_index] = [column_name, index+2,
                                                        "Warnings:When the 'Alternative' column is not empty, sign column should not be empty, it may cause errors."]
                            next_index += 1
    
                    for route_id_column in ["Route Id", "route id", "Route id", "route Id"]:
                        if route_id_column in df.columns:
                            if not pd.isna(row[route_id_column]):
                                error_df.loc[next_index] = [route_id_column, index+2,
                                                            f"When the 'Alternative' column is not empty, the '{route_id_column}' should be empty"]
                                next_index += 1
                # Here, we filter the dataframe for rows where the Alternative column is empty or has a "-"
                filtered_df = df[(df[column_name] == "-") | df[column_name].isna()]
    
                # For these rows, check if there is a Route Id column in the main DataFrame
                for route_id_column in ["Route Id", "route id", "Route id", "route Id"]:
                    if route_id_column in df.columns:
                        for index, row in filtered_df.iterrows():
                            expected_route_id = f'{row[df.columns[3]]}-{row[df.columns[4]]}-{row[df.columns[6]]}-{row[df.columns[7]]}-{row[df.columns[12]]}'
                            if row[route_id_column] != expected_route_id:
                                error_df.loc[next_index] = [route_id_column, index+2,
                                    f"The 'Route Id' column should follow the format: 'column index 3-column index 4-column index 6-column index 7-column index 12'. The expected value is {expected_route_id}"]
                                next_index += 1
                        break
                    else:
                        if route_id_column == "route Id":  # After all checks, the column was not found.
                            for index, _ in filtered_df.iterrows():
                                error_df.loc[next_index] = [column_name, index+2,
                                    "When the 'Alternative' column is empty, you should have a 'Route Id' column with the format: sign-direction-start stop ID-end stop ID-distance"]
                                next_index += 1
    
                # Else case, where the '-' sign should be deleted
                error_rows = df[(df[column_name] == "-") | (df[column_name] == -1)].index.tolist()
                for row in error_rows:
                    error_df.loc[next_index] = [column_name, row+2,
                                                "Try to delete the '-' sign in the 'Alternative' column and keep the value empty."]
                    next_index += 1
            elif column_index == 12:  # Check if value is numeric
                for idx, val in df[column_name].items():
                    if not str(val).replace('.', '', 1).isdigit():
                        error_df.loc[next_index] = [column_index, idx+2, "Invalid value. Value must be numeric."]
                        next_index += 1
            # Check if value is in allowed set
            elif column_index == 15:
                # Convert to string and sort the characters
                df[column_name] = df[column_name].astype(str).apply(lambda x: "".join(sorted(x)))
    
                # Generate allowed values. Create combinations of 1-7, sort and remove duplicates
                from itertools import combinations, chain
    
                valid_days = []
                for i in range(1, 8):
                    # create combinations for each length
                    combos = combinations(range(1, 8), i)
                    # add each sorted combination to valid_days
                    for c in combos:
                        valid_days.append("".join(sorted(map(str, c))))
    
                # Check if the sorted days are in the allowed set
                if not df[column_name].isin(valid_days).all():
                    error_rows = df[~df[column_name].isin(valid_days)].index.tolist()
                    for row in error_rows:
                        error_df.loc[next_index] = [column_index, row+2, "Invalid day. Only combinations of 1-7 are allowed."]
                        next_index += 1
        # Create new column that is the combination of columns index 3, 4, and 5
        df["combined"] = df[df.columns[3]].astype(str) + " " + df[df.columns[4]].astype(str) + " " + df[df.columns[5]].astype(str)
        # Group by the combined column and check if column index 6 and 7 have the same value for each group
        grouped = df.groupby("combined")
        for name, group in grouped:
            if len(group[df.columns[6]].unique()) > 1 or len(group[df.columns[7]].unique()) > 1:
                error_df.loc[next_index] = ['Start stop/end stop', name,
                    "All unique combinations of columns sign+direction+alternative should have the same values in the start stop and end stop"]
                next_index += 1
    
        # Drop the combined column as it is not needed anymore
        df = df.drop(columns=["combined"])
    
        return error_df
    
    def validate_trips_with_stoptimes(excel_sheet):
        wb = load_workbook(excel_sheet, read_only=True)  # open an Excel file and return a workbook
        if 'Trips' not in wb.sheetnames:
            data = {'Trips': ['Trips sheet not exists']}
            df = pd.DataFrame(data)
            return df
        if 'StopTimes' not in wb.sheetnames:
            data = {'StopTimes': ['StopTimes sheet not exists']}
            df = pd.DataFrame(data)
            return df
    
        trips = pd.read_excel(excel_sheet, sheet_name='Trips')
        stoptimes = pd.read_excel(excel_sheet, sheet_name='StopTimes')
    
        num_rows = stoptimes.shape[0]
        if num_rows<2:
            data = {'StopTimes': ['StopTimes sheet not exists']}
            df = pd.DataFrame(data)
            return df
        # Check for the column name and assign it to a variable
        if "Trip Id" in stoptimes.columns:
            trip_id_col = "Trip Id"
        elif "TRIP_ID" in stoptimes.columns:
            trip_id_col = "TRIP_ID"
        else:
            raise ValueError("Neither 'Trip Id' nor 'TRIP_ID' found in the StopTimes sheet")
    
        # Prepare error data frame
        error_df = pd.DataFrame(columns=['Id', 'Error'])
    
        # Check if all Ids in trips dataframe are in the stoptimes dataframe
        id_check_mask = trips['Id'].isin(stoptimes[trip_id_col])
        if not all(id_check_mask):
            temp_df = trips.loc[~id_check_mask, ['Id']].copy()
            temp_df['Error'] = "Trip ID does not appear in the stoptimes tab."
            error_df = pd.concat([error_df, temp_df], ignore_index=True)
    
        # Merge trips with stoptimes on Ids
        merged_df = pd.merge(trips, stoptimes, how='inner', left_on='Id', right_on=trip_id_col)
    
        # Check for the correct column names and assign them to variables
        sequence_col = 'Sequence' if 'Sequence' in merged_df.columns else 'SEQUENCE'
        time_col = 'Time' if 'Time' in merged_df.columns else 'TIME'
        origin_stop_id_col = 'Origin Stop id' if 'Origin Stop id' in merged_df.columns else 'Origin Stop Id'
        point_id_col = 'Point Id' if 'Point Id' in merged_df.columns else 'POINT_ID'
    
        # Convert sequence to int
        merged_df[sequence_col] = merged_df[sequence_col].astype(int)
    
        # Get the first and last sequence Time for each trip
        first_sequence_df = merged_df.loc[merged_df.groupby('Id')[sequence_col].idxmin()]
        last_sequence_df = merged_df.loc[merged_df.groupby('Id')[sequence_col].idxmax()]
    
        # Check if Departure in trips matches Time at first sequence in stoptimes
        if not all(first_sequence_df['Departure'] == first_sequence_df[time_col]):
            temp_df = first_sequence_df.loc[first_sequence_df['Departure'] != first_sequence_df[time_col], ['Id']].copy()
            temp_df['Error'] = "Departure time does not match the time at first sequence in stoptimes."
            error_df = pd.concat([error_df, temp_df], ignore_index=True)
    
        # Check if Arrival in trips matches Time at last sequence in stoptimes
        if not all(last_sequence_df['Arrival'] == last_sequence_df[time_col]):
            temp_df = last_sequence_df.loc[last_sequence_df['Arrival'] != last_sequence_df[time_col], ['Id']].copy()
            temp_df['Error'] = "Arrival time does not match the time at last sequence in stoptimes."
            error_df = pd.concat([error_df, temp_df], ignore_index=True)
    
        # Check if Origin Stop id in trips matches Point Id at first sequence in stoptimes
        if not all(first_sequence_df[origin_stop_id_col] == first_sequence_df[point_id_col]):
            temp_df = first_sequence_df.loc[first_sequence_df[origin_stop_id_col] != first_sequence_df[point_id_col], ['Id']].copy()
            temp_df['Error'] = f"{origin_stop_id_col} does not match the {point_id_col} at first sequence in stoptimes."
            error_df = pd.concat([error_df, temp_df], ignore_index=True)
    
        # Check if Destination Stop Id in trips matches Point Id at last sequence in stoptimes
        if not all(last_sequence_df['Destination Stop Id'] == last_sequence_df[point_id_col]):
            temp_df = last_sequence_df.loc[last_sequence_df['Destination Stop Id'] != last_sequence_df[point_id_col], ['Id']].copy()
            temp_df['Error'] = "Destination Stop Id does not match the Point Id at last sequence in stoptimes."
            error_df = pd.concat([error_df, temp_df], ignore_index=True)
    
        return error_df
    
    def validate_trips_with_places(excel_sheet):
        wb = load_workbook(excel_sheet, read_only=True)  # open an Excel file and return a workbook
        if 'Trips' not in wb.sheetnames:
            data = {'Trips': [
                'Trips sheet not exists']}
            df = pd.DataFrame(data)
            return df
        if 'Places' not in wb.sheetnames:
            data = {'Places': [
                'Places sheet not exists']}
            df = pd.DataFrame(data)
            return df
        trips = pd.read_excel(excel_sheet, sheet_name='Trips')
        places = pd.read_excel(excel_sheet, sheet_name='Places')
        # Define the columns for stop ids in the Trips dataframe
        stop_id_cols = [trips.columns[6], trips.columns[7]]
    
        # Get the unique stop ids from the Trips dataframe
        trip_stop_ids = pd.concat([trips[col] for col in stop_id_cols]).unique()
    
        # Get the unique ids from the Places dataframe
        place_ids = places['Id'].unique()
    
        # Find the stop ids from Trips that are missing in Places
        missing_ids = [id for id in trip_stop_ids if id not in place_ids]
    
        # Create a dataframe to hold the issues
        issues_df = pd.DataFrame(columns=['Issue Column', 'Row Number', 'Error Message'])
    
        # If there are missing ids, add details to the issues dataframe
        if missing_ids:
            for col in stop_id_cols:
                for index, row in trips.iterrows():
                    if row[col] in missing_ids:
                        # Create a temporary dataframe to hold the issue details
                        temp_df = pd.DataFrame({
                            'Issue Column': [col],
                            'Row Number': [index],
                            'Error Message': [f'Stop id {row[col]} does not exist in the Places tab.']
                        })
    
                        # Append the temporary dataframe to the issues dataframe
                        issues_df = pd.concat([issues_df, temp_df], ignore_index=True)
    
        return issues_df
    
    def validate_trips_with_veihcles(excel_sheet):
        wb = load_workbook(excel_sheet, read_only=True)  # open an Excel file and return a workbook
        if 'Trips' not in wb.sheetnames:
            data = {'Trips': [
                'Trips sheet not exists']}
            df = pd.DataFrame(data)
            return df
        if 'VehicleTypes' not in wb.sheetnames:
            data = {'VehicleTypes': [
                'VehicleTypes sheet not exists']}
            df = pd.DataFrame(data)
            return df
        Trips = pd.read_excel(excel_sheet, sheet_name='Trips',dtype='str')
        VehicleTypes = pd.read_excel(excel_sheet, sheet_name='VehicleTypes',dtype='str')
        # Step 1: Identify the correct column in the Trips dataframe
        vehicle_col = None
        for col in ["Vehicle Type Ids", "Vehicle Type Id", "קוד סוג רכב"]:
            if col in Trips.columns:
                vehicle_col = col
                break
        if vehicle_col is None:
            return pd.DataFrame({'Error Column': [], 'Row Number': [], 'Error Message': []})
    
        # Step 2: Split by '|' and get the unique list of vehicle types in Trips
        vehicle_types_in_trips = set()
        for i, row in Trips.iterrows():
            vehicle_types = row[vehicle_col].split("|")
            vehicle_types_in_trips.update(vehicle_types)
        # Step 3: Compare with the vehicle types in VehicleTypes
        vehicle_types_in_vehicleTypes = set(VehicleTypes.iloc[:,0])
        missing_vehicle_types = vehicle_types_in_trips - vehicle_types_in_vehicleTypes
        # Generate the error DataFrame
        errors = []
        for i, row in Trips.iterrows():
            vehicle_types = str(row[vehicle_col].split("|"))
            for vehicle_type in vehicle_types:
                if vehicle_type in missing_vehicle_types:
                    errors.append({
                        'Error Column': vehicle_col,
                        'Row Number': i+2,
                        'Error Message': f'vehicle type "{vehicle_type}" appears in the Trips tab but not in the VehicleTypes tab '
                    })
        return pd.DataFrame(errors)
    
    def general():
        data = {'General Comments': ['If you have tabs that are not: Places, stoptimes, trips or VehicleTypes, consider to remove them since it can cause errors', 'Before importing the dataset, make sure that the feature flag is on', 'Please notice that when you uploading a new dataset, you will be overwriting the existing stop catalog in the project']}
        df = pd.DataFrame(data)
        return df

    def write_excel(table_array):
        time_0 = time()
        logger.write('\nWriting excel output')
        output = BytesIO()
        excel = ExcelWriter(output, engine='xlsxwriter')
        header_format = excel.book.add_format({'bold': True, 'text_wrap': False, 'align': 'left'})
        for i in table_array:
            table_array[i].to_excel(excel, sheet_name=i, merge_cells=False, freeze_panes=[1, 0], index=False)
            for col_num, value in enumerate(table_array[i].columns.values):
                excel.sheets[i].write(0, col_num, value, header_format)
                excel.sheets[i].set_column(col_num, col_num,
                                           max(table_array[i][value].astype(str).str.len().max(), len(value) + 2))
        excel.close()
        output.seek(0)
        return output


    main()
